from openpyxl import load_workbook

file = "D:/workdir/12stadb/uniprotkb_PTM_Processing_palmitoylation_2025_02_05.xlsx"

wb = load_workbook(file)

st = wb.worksheets[0]

outfile = open("D:/workdir/12stadb/palmitoylation.txt","a")
outfile.write("Gene\n")
for data in st.iter_rows(2,st.max_row):
	gene = data[2].value.split("_")[0].strip()

	outfile.write(f"{gene}\n")